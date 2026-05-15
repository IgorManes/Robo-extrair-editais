import os
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CABECALHOS = [
    "ORGAO", "LINK", "CAMINHO", "NOME_EDITAL",
    "DATA_EXECUCAO", "STATUS_EXECUCAO", "OBSERVACOES"
]

LARGURAS_COLUNAS = [18, 45, 55, 60, 16, 18, 50]
COR_CABECALHO    = "1F4E79"
COR_SUCESSO      = "C6EFCE"
COR_FALHA        = "FFC7CE"


def _aplicar_estilo_cabecalho(cell):
    cell.font      = Font(bold=True, color="FFFFFF", size=11)
    cell.fill      = PatternFill("solid", fgColor=COR_CABECALHO)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = Border(bottom=Side(style="thin", color="FFFFFF"))


def gerar_relatorio(registros: list[dict], pasta_data: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title     = "Relatório"
    ws.row_dimensions[1].height = 20

    ws.append(CABECALHOS)
    for idx, cell in enumerate(ws[1], start=1):
        _aplicar_estilo_cabecalho(cell)
        ws.column_dimensions[cell.column_letter].width = LARGURAS_COLUNAS[idx - 1]

    for reg in registros:
        linha = [reg.get(col, "") for col in CABECALHOS]
        ws.append(linha)

        row_num = ws.max_row
        cor     = COR_SUCESSO if reg.get("STATUS_EXECUCAO") == "Sucesso" else COR_FALHA
        fill    = PatternFill("solid", fgColor=cor)

        for cell in ws[row_num]:
            cell.fill      = fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    ws.freeze_panes = "A2"

    caminho = os.path.join(pasta_data, f"relatorio_{date.today()}.xlsx")
    wb.save(caminho)
    return caminho